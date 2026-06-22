import flet as ft
import os, datetime
import uuid
from styles import input_style, drop_style, datatable_style, stat_style, login_style, switch_style
from components.components import MyButton
from utils import (
    ACCESS_TOKEN, TENANT_ID, ROLE, MAIN_COLOR, BG_COLOR, resource_path, format_milliers_fr, CARD_BG, BORDER_COLOR, SHADOW_COLOR
)
import asyncio, threading
from services.async_function import supabase_request_async
from services.supabase_client import supabase_client, supabase_admin

DEFAULT_IMAGE = "https://hojfmjmrhtsvgfzynelr.supabase.co/storage/v1/object/public/images/no%20image.jpeg"


class Products(ft.Container):
    def __init__(self, cp: object):
        super().__init__(expand=True)
        self.cp = cp

        # Récupération des données d'authentification
        self.tenant_id = self.cp.page.client_storage.get(TENANT_ID)
        self.access_token = self.cp.page.client_storage.get(ACCESS_TOKEN)
        self.user_role = self.cp.page.client_storage.get(ROLE)

        # Liste des produits
        self.liste_des_produits: list = []
        self.selected_image_path = None

        # Statistiques
        self.stat_total_produits = ft.Text("0", size=22, font_family="PEB")
        self.stat_valeur_stock = ft.Text("0", size=22, font_family="PEB")
        self.stat_categories = ft.Text("0", size=22, font_family="PEB")

        self.loader = ft.ProgressRing(width=20, height=20, stroke_width=2, color=MAIN_COLOR, visible=False)

        # Recherche
        self.search_field = ft.TextField(
            **input_style,
            width=350,
            label="Rechercher",
            prefix_icon=ft.Icons.SEARCH_ROUNDED,
            on_change=self.filter_products_locally
        )

        # Tableau principal avec une colonne supplémentaire pour le transfert
        self.data_table = ft.DataTable(
            **datatable_style,
            columns=[
                ft.DataColumn(ft.Row(
                    [ft.Image(resource_path('assets/icons/grey/shopping-cart.svg'), width=18, height=18),
                     ft.Text("Désignation")])),
                ft.DataColumn(ft.Row(
                    [ft.Image(resource_path('assets/icons/grey/tag.svg'), width=18, height=18), ft.Text("Catégorie")])),
                ft.DataColumn(ft.Row([ft.Image(resource_path('assets/icons/grey/badge-cent.svg'), width=18, height=18),
                                      ft.Text("Prix")])),
                ft.DataColumn(ft.Row([ft.Image(resource_path('assets/icons/grey/store.svg'), width=18, height=18),
                                      ft.Text("Stock boutique")])),
                ft.DataColumn(ft.Row([ft.Image(resource_path('assets/icons/grey/sigma.svg'), width=18, height=18),
                                      ft.Text("Stock tampon")])),
                ft.DataColumn(ft.Row(
                    [ft.Image(resource_path('assets/icons/grey/circle-dollar-sign.svg'), width=18, height=18),
                     ft.Text("Valeur")])),
                ft.DataColumn(ft.Text("")),  # modification (crayon)
            ],
        )

        self.table_container = ft.ListView(controls=[self.data_table], expand=True, spacing=10)

        # Champs du formulaire produit (inchangés)
        self.p_name = ft.TextField(**input_style, label="Désignation du produit", width=440,
                                   capitalization=ft.TextCapitalization.WORDS)
        self.p_category = ft.Dropdown(**drop_style, label="Catégorie", width=440, options=[])
        self.p_price = ft.TextField(**input_style, label="Prix vente", width=170, input_filter=ft.NumbersOnlyInputFilter(),
                                    text_align=ft.TextAlign.RIGHT)
        self.p_price_buy = ft.TextField(**input_style, label="Prix achat", width=170,
                                        input_filter=ft.NumbersOnlyInputFilter(), text_align=ft.TextAlign.RIGHT)
        self.p_stock = ft.TextField(**input_style, label="Stock Initial", width=170, value="0",
                                    input_filter=ft.NumbersOnlyInputFilter(), text_align=ft.TextAlign.RIGHT,
                                    disabled=True)
        self.p_image_name = ft.TextField(**input_style, label="Image sélectionnée", width=300, read_only=True,
                                         hint_text="Aucune image choisie (Image par défaut)")

        self.p_new_category = ft.TextField(
            **input_style, width=440, capitalization=ft.TextCapitalization.WORDS, disabled=True,
            label="Nouvelle catégorie",
        )
        self.p_switch = ft.Switch(
            **switch_style, scale=0.7,
            on_change=self.on_change_category
        )

        #champs du formulaire de modification
        self.selected_product_id = 0
        self.edit_designation_field = ft.TextField(**input_style, label="Désignation du produit", width=440,
                                   capitalization=ft.TextCapitalization.WORDS)
        self.edit_price_field = ft.TextField(**input_style, label="Prix", width=170, input_filter=ft.NumbersOnlyInputFilter(),
                                    text_align=ft.TextAlign.RIGHT)
        self.edit_price_buy_field = ft.TextField(**input_style, label="Prix achat", width=170,
                                        input_filter=ft.NumbersOnlyInputFilter(), text_align=ft.TextAlign.RIGHT)
        self.edit_type_field = ft.TextField(**input_style, label="Désignation du produit", width=440,
                                   capitalization=ft.TextCapitalization.WORDS)
        # ================= ZONE EDITION DE PRODUIT =================
        # Image par défaut pour la prévisualisation dans la fenêtre de modification
        self.fp = ft.FilePicker(on_result=self.on_file_picker_result)
        self.cp.page.overlay.append(self.fp)
        self.edit_image = ft.Image(
            src=DEFAULT_IMAGE,
            width=100,
            height=100,
            fit=ft.ImageFit.COVER,
            border_radius=ft.border_radius.all(50)
        )

        # Bouton d'upload positionné par-dessus ou à côté de l'image
        self.upload_btn = ft.IconButton(
            icon=ft.Icons.ADD_A_PHOTO_ROUNDED,
            icon_color=MAIN_COLOR,
            bgcolor=ft.Colors.WHITE,
            icon_size=20,
            tooltip="Changer l'image",
            right=0,
            bottom=0,
            on_click=lambda _: self.fp.pick_files(allow_multiple=False, file_type=ft.FilePickerFileType.IMAGE)
        )

        # Variable pour stocker le produit en cours de modification
        self.selected_product_id = None

        # File pickers
        self.excel_picker = ft.FilePicker(on_result=self.on_excel_file_picked)
        self.image_picker = ft.FilePicker(on_result=self.on_image_file_picked)
        self.cp.page.overlay.extend([self.excel_picker, self.image_picker])

        # Bouton pour générer et télécharger le modèle Excel
        self.btn_download_template = MyButton(
            title="Modèle Excel",
            icon=ft.Icons.DOWNLOAD_ROUNDED,
            click=self.generate_excel_template
        )

        # Le bouton d'import existant appellera maintenant le choix de l'emplacement du stock
        self.btn_import_excel = MyButton(
            title="Importer Excel",
            icon=ft.Icons.UPLOAD_FILE_ROUNDED,
            click=self.open_import_destination_modal
        )

        # Variable de classe à initialiser dans __init__
        self.selected_stock_destination = "products"  # "products" = Boutique, "stock_tampons" = Tampon

        # ================= ZONE APPROVISIONNEMENT STOCK TAMPON =================
        self.tampon_dropdown_produits = ft.Dropdown(**drop_style, label="Sélectionner le produit", width=300, menu_height=200)
        self.tampon_qty_field = ft.TextField(
            **login_style, label="Quantité", hint_text="Ex: 50", width=120,
            value="1", input_filter=ft.NumbersOnlyInputFilter()
        )
        self.tampon_panier_table = ft.DataTable(
            **datatable_style,
            columns=[
                ft.DataColumn(ft.Text("Désignation")),
                ft.DataColumn(ft.Text("Quantité")),
                ft.DataColumn(ft.Text("Action")),
            ]
        )
        # Liste locale pour stocker les éléments du panier
        self.tampon_panier_items = []


        # Liste Python locale pour stocker les éléments du panier avant validation
        self.tampon_panier_items = []

        has_admin_rights = self.user_role in ["admin", "manager"]

        self.content = ft.Column(
            controls=[
                ft.Row(
                    controls=[
                        ft.Column([ft.Text("Gestion du Catalogue Produits", size=26, font_family="PEB")], spacing=4),
                        ft.Row([self.loader], spacing=15, vertical_alignment=ft.CrossAxisAlignment.CENTER)
                    ],
                    alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                ),
                ft.Divider(height=15, color=ft.Colors.TRANSPARENT),
                self.build_stats_row(),
                ft.Divider(height=15, color=ft.Colors.TRANSPARENT),
                ft.Container(
                    **stat_style, expand=True,
                    content=ft.Column(
                        expand=True,
                        controls=[
                            ft.Row(
                                controls=[
                                    self.search_field,
                                    ft.Row(
                                        controls=[
                                            # Suppression du bouton Transférer Stock
                                            MyButton(
                                                "Ajouter un produit", resource_path("assets/icons/white/user-plus.svg"),
                                                     click=self.open_add_product_container
                                            ),
                                            MyButton(
                                                title="Modèle Excel",
                                                icon=ft.Icons.DOWNLOAD_ROUNDED,
                                                click=self.generate_excel_template
                                            ),
                                            MyButton("Importer via excel", resource_path("assets/icons/white/sheet.svg"),
                                                     click=self.open_import_destination_modal),
                                            MyButton(
                                                "Entrées Stock Tampon",
                                                resource_path("assets/icons/white/plus.svg"),
                                                # Ajuste le chemin de ton icône
                                                lambda e: self.run_async_in_thread(self.open_entree_tampon_window(e))
                                            )
                                        ],
                                        visible=has_admin_rights
                                    )
                                ],
                                alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                            ),
                            ft.Divider(height=10, color=ft.Colors.TRANSPARENT),
                            self.table_container
                        ]
                    )
                ),
            ],
            spacing=10
        )
        self.on_mount()

    def on_change_category(self, e):
        # En Flet, pour réactiver un champ, on passe .disabled à False
        if self.p_switch.value:
            self.p_new_category.disabled = False
        else:
            self.p_new_category.disabled = True
            self.p_new_category.value = ""

        # Mise à jour ciblée du container ou de la page sécurisée
        try:
            self.p_new_category.update()
            self.p_switch.update()
        except Exception:
            if self.cp.page:
                self.cp.page.update()

    def build_stats_row(self):
        def create_stat_card(title, text_control, icon):
            return ft.Container(
                **stat_style, width=285,
                content=ft.Column(
                    controls=[
                        ft.Row(
                            controls=[
                                ft.Text(title, size=13, font_family="PPR", color='grey'),
                                ft.Image(resource_path(icon), width=18, height=18),
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                        ),
                        text_control
                    ],
                ),
            )
        return ft.Row([
            create_stat_card("Articles Référencés", self.stat_total_produits, "assets/icons/grey/shopping-cart.svg"),
            create_stat_card("Valeur du Stock", self.stat_valeur_stock, "assets/icons/grey/badge-cent.svg"),
            create_stat_card("Catégories Actives", self.stat_categories, "assets/icons/grey/tag.svg"),
        ], spacing=15, wrap=True)

    @staticmethod
    def run_async_in_thread(coro):
        def runner():
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(coro)
            loop.close()
        threading.Thread(target=runner).start()

    def on_mount(self):
        self.run_async_in_thread(self.load_products_data())

    async def load_products_data(self):
        self.loader.visible = True
        self.update()

        try:
            params = {'select': '*', 'tenant_id': f'eq.{self.tenant_id}', 'order': 'designation.asc'}
            response = await supabase_request_async(
                access_token=self.access_token, tenant_id=self.tenant_id,
                table_name="v_produits_avec_tampon", method="GET", params=params
            )
            self.liste_des_produits = response if (response and isinstance(response, list)) else []
        except Exception as ex:
            print(f"Erreur de chargement des produits : {ex}")
            self.liste_des_produits = []

        self.calculate_metrics()
        self.render_products_table(self.liste_des_produits)
        self.loader.visible = False
        self.cp.page.update()

    def calculate_metrics(self):
        total_items = len(self.liste_des_produits)
        valeur_totale = 0
        categories_uniques = set()

        for p in self.liste_des_produits:
            stock = int(p.get("stock_reel", 0))
            prix_vente = float(p.get("price_sell", 0))
            valeur_totale += (stock * prix_vente)
            if p.get("product_type"):
                categories_uniques.add(p.get("product_type"))

        self.p_category.options.clear()
        for category in categories_uniques:
            self.p_category.options.append(ft.dropdown.Option(key=category, text=category))

        self.stat_total_produits.value = str(total_items)
        self.stat_valeur_stock.value = f"{format_milliers_fr(valeur_totale)}"
        self.stat_categories.value = str(len(categories_uniques))

    def render_products_table(self, products_list):
        self.data_table.rows.clear()
        can_edit = self.user_role in ["admin", "manager"]

        for prod in products_list:
            stock_actuel = int(prod.get("stock_reel", 0))
            p_vente = float(prod.get("price_sell", 0))
            valeur = int(stock_actuel * p_vente)
            stock_tampon = int(prod.get("stock_tampon", 0))
            row_color = "red" if stock_actuel <= 0 else "black"

            # Bouton Modifier
            if can_edit:
                edit_btn = ft.Container(
                    content=ft.Image(resource_path("assets/icons/grey/pen.svg"), width=18, height=18),
                    on_click=lambda e, p=prod: self.open_edit_window(p),
                    data=prod
                )
                transfer_btn = ft.Container(
                    content=ft.Image(resource_path("assets/icons/grey/folder-sync.svg"), width=18, height=18),
                    on_click=lambda e, p=prod: self.open_transfer_modal(p),
                    data=prod, tooltip="Transférer du stock tampon vers boutique",
                )
            else:
                edit_btn = ft.Text("")
                transfer_btn = ft.Text("")
            # Bouton Transférer (visible pour admin/manager)


            row_item = ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(prod.get("designation", "Sans désignation"))),
                    ft.DataCell(ft.Text(prod.get("product_type", "Général"))),
                    ft.DataCell(ft.Text(f"{format_milliers_fr(p_vente)}")),
                    ft.DataCell(ft.Text(f"{format_milliers_fr(stock_actuel)}", color=row_color)),
                    ft.DataCell(ft.Text(f"{format_milliers_fr(stock_tampon)}", color=row_color)),
                    ft.DataCell(ft.Text(f"{format_milliers_fr(valeur)}", color=row_color)),
                    ft.DataCell(
                        ft.Row(
                            controls=[
                                edit_btn, transfer_btn
                            ]
                        )
                    ),
                ]
            )
            self.data_table.rows.append(row_item)
        self.cp.page.update()

    # --- Transfert individuel ---
    def open_transfer_modal(self, product_data):
        """Ouvre un modal pour transférer une quantité du tampon vers la boutique"""
        product_id = product_data.get("produit_id")
        designation = product_data.get("designation", "")
        stock_boutique = int(product_data.get("stock_reel", 0))
        stock_tampon = int(product_data.get("stock_tampon", 0))

        # Champs du modal
        name_text = ft.Text(f"Produit : {designation}", size=18, font_family="PEB")
        stock_info = ft.Text(f"Stock boutique : {stock_boutique}  |  Stock tampon : {stock_tampon}", size=16, font_family="PPM")
        qty_field = ft.TextField(
            **input_style,
            label="Quantité à transférer",
            width=200,
            value="1",
            input_filter=ft.NumbersOnlyInputFilter(),
            text_align=ft.TextAlign.RIGHT,
        )

        # Bouton Valider
        def on_validate(e):
            try:
                qty = int(qty_field.value)
                if qty <= 0:
                    self.cp.show_alert("La quantité doit être positive.", ft.Icons.WARNING, ft.Colors.ORANGE)
                    return
                if qty > stock_tampon:
                    self.cp.show_alert(f"Quantité demandée ({qty}) dépasse le stock tampon ({stock_tampon}).", ft.Icons.ERROR, ft.Colors.RED)
                    return
                # Fermer le modal
                self.cp.hide_container(self.cp.st_container)
                # Lancer le transfert
                self.run_async_in_thread(self.execute_transfer(product_id, qty))
            except ValueError:
                self.cp.show_alert("Veuillez entrer un nombre valide.", ft.Icons.WARNING, ft.Colors.ORANGE)

        # Construction du modal
        self.cp.st_form.content = None
        self.cp.st_container.content.height = 320
        self.cp.st_container.content.width = 400
        self.cp.st_form.content = ft.Column(
            controls=[
                ft.Row([
                    ft.Text("Transfert de stock", size=22, font_family="PEB"),
                    ft.Container(
                        content=ft.Image(
                            resource_path('assets/icons/black/x.svg'), width=18, height=18
                        ),
                        on_click=lambda _: self.cp.hide_container(self.cp.st_container)
                    ),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Divider(height=10, thickness=1),
                name_text,
                stock_info,
                ft.Divider(height=10, color="transparent"),
                qty_field,
                ft.Row([
                    MyButton("Valider le transfert", resource_path("assets/icons/white/check.svg"), on_validate)
                ], alignment=ft.MainAxisAlignment.END)
            ],
            spacing=15,
        )
        self.cp.show_container(self.cp.st_container)
        self.cp.page.update()

    async def execute_transfer(self, product_id: str, qty: int):
        """Effectue la mise à jour des stocks en boutique et tampon, puis consigne l'entrée"""
        if not product_id:
            self.cp.show_alert("Identifiant produit manquant.", ft.Icons.ERROR, ft.Colors.RED)
            return

        self.loader.visible = True
        self.cp.page.update()

        try:
            # 1. Récupérer les stocks actuels via la vue
            params = {'select': '*', 'produit_id': f'eq.{product_id}'}
            result = await supabase_request_async(
                access_token=self.access_token,
                tenant_id=self.tenant_id,
                table_name="v_produits_avec_tampon",
                method="GET",
                params=params
            )
            if not result or len(result) == 0:
                self.cp.show_alert("Produit introuvable.", ft.Icons.ERROR, ft.Colors.RED)
                return

            data = result[0]

            # Récupération sécurisée des stocks (convertir None en 0)
            stock_boutique = int(data.get("stock_reel") or 0)
            stock_tampon = int(data.get("stock_tampon") or 0)
            prix_unitaire = float(data.get("price") or 0.0)  # Récupération du prix pour l'historique d'entrée

            # 2. Validation des quantités
            if qty <= 0:
                self.cp.show_alert("La quantité doit être positive.", ft.Icons.WARNING, ft.Colors.ORANGE)
                return
            if qty > stock_tampon:
                self.cp.show_alert(f"Stock tampon insuffisant. Disponible : {stock_tampon}.", ft.Icons.ERROR,
                                   ft.Colors.RED)
                return

            # 3. Calcul des nouveaux stocks
            new_stock_boutique = stock_boutique + qty
            new_stock_tampon = stock_tampon - qty

            # 4. Mise à jour de la table products (stock boutique)
            update_payload = {"stock": new_stock_boutique}
            await supabase_request_async(
                access_token=self.access_token,
                tenant_id=self.tenant_id,
                table_name="products",
                method="PATCH",
                params={"id": f"eq.{int(product_id)}"},
                data=update_payload
            )

            # 5. Mise à jour de la table stock_tampons (stock tampon)
            update_payload_tampon = {"stock": new_stock_tampon}
            await supabase_request_async(
                access_token=self.access_token,
                tenant_id=self.tenant_id,
                table_name="stock_tampons",
                method="PATCH",
                params={"id": f"eq.{int(product_id)}"},
                data=update_payload_tampon
            )

            # 6. NOUVEAU : Consignation du mouvement dans entrees_details pour la RPC de clôture
            # Note : La colonne 'date' prendra automatiquement la date du jour grâce au DEFAULT SQL configuré,
            # mais envoyer le payload ainsi permet de structurer proprement la ligne.
            payload_entree = {
                "tenant_id": self.tenant_id,
                "id_produit": int(product_id),
                "qte": int(qty),
                "prix": prix_unitaire,
                "numero": f"TRF-{datetime.date.today().strftime('%Y%m%d')}"
                # Génère une référence automatique de type TRF-20260620
            }

            await supabase_request_async(
                access_token=self.access_token,
                tenant_id=self.tenant_id,
                table_name="entrees_details",
                method="POST",
                data=payload_entree
            )

            # 7. Recharger l'affichage local et notifier l'utilisateur
            await self.load_products_data()
            self.cp.show_alert(
                f"Transfert de {qty} unité(s) effectué et comptabilisé avec succès !",
                ft.Icons.CHECK_CIRCLE,
                ft.Colors.GREEN
            )

        except Exception as ex:
            print(f"Erreur lors du transfert : {ex}")
            self.cp.show_alert("Erreur lors du transfert en base.", ft.Icons.ERROR, ft.Colors.RED)
        finally:
            self.loader.visible = False
            self.cp.page.update()

    # --- Filtre de recherche ---
    def filter_products_locally(self, e):
        query = self.search_field.value.lower().strip()
        if not query:
            self.render_products_table(self.liste_des_produits)
            return
        filtered = [
            p for p in self.liste_des_produits
            if query in p.get("designation", "").lower() or query in p.get("product_type", "").lower()
        ]
        self.render_products_table(filtered)

    # --- Gestion des images ---
    def on_image_file_picked(self, e: ft.FilePickerResultEvent):
        if e.files and len(e.files) > 0:
            self.selected_image_path = e.files[0].path
            self.p_image_name.value = os.path.basename(self.selected_image_path)
            self.p_image_name.update()

    # --- Ajout de produit ---
    def open_add_product_container(self, e):
        self.p_name.value = ""
        self.p_name.disabled = False
        self.p_category.value = "Général"
        self.p_price.value = ""
        self.p_stock.value = "0"
        self.p_image_name.value = ""
        self.selected_image_path = None

        self.cp.st_container.content.height = 620
        self.cp.st_container.content.width = 500
        self.cp.st_form.content = None

        self.cp.st_form.content = ft.Column(
            controls=[
                ft.Row([
                    ft.Text("Nouveau Produit", size=22, font_family="PEB"),
                    ft.Container(
                        content=ft.Image(resource_path('assets/icons/black/x.svg'), width=18, height=18),
                        on_click=lambda e: self.cp.hide_container(self.cp.st_container)
                    )
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                ft.Divider(height=5, color="grey200"),
                ft.Divider(height=10, color="transparent"),
                ft.Column(
                    spacing=10, scroll=ft.ScrollMode.ADAPTIVE,
                    controls=[
                        self.p_name, self.p_category,
                        ft.Row(
                            [
                                ft.Text("Activer nouvelle catégorie", size=14, font_family="PPM"),
                                self.p_switch
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN
                        ), self.p_new_category,
                        self.p_price, self.p_price_buy, self.p_stock,
                        ft.Row([
                            self.p_image_name,
                            ft.IconButton(
                                icon=ft.Icons.IMAGE_SEARCH_ROUNDED, icon_color=MAIN_COLOR, tooltip="Choisir une image",
                                on_click=lambda _: self.image_picker.pick_files(
                                    allowed_extensions=["png", "jpg", "jpeg", "webp", "svg"])
                            )
                        ], alignment=ft.MainAxisAlignment.START),
                        ft.Divider(height=15, color="transparent"),
                        ft.Row([
                            MyButton("Enregistrer l'article", resource_path("assets/icons/white/user-check.svg"),
                                     click=lambda e: self.run_async_in_thread(self.save_new_product(e)))
                        ], alignment=ft.MainAxisAlignment.END)
                    ]
                )
            ]
        )
        self.cp.show_container(self.cp.st_container)
        self.cp.update()

    async def save_new_product(self, e):
        if not self.p_name.value or not self.p_price.value:
            self.cp.show_alert("Veuillez remplir les champs obligatoires !", ft.Icons.ERROR_OUTLINED, ft.Colors.RED)
            return

        self.loader.visible = True
        self.update()
        final_image_url = DEFAULT_IMAGE

        if self.selected_image_path and os.path.exists(self.selected_image_path):
            try:
                file_extension = os.path.splitext(self.selected_image_path)[1]
                unique_filename = f"{self.tenant_id}/{uuid.uuid4()}{file_extension}"
                with open(self.selected_image_path, "rb") as image_file:
                    file_bytes = image_file.read()

                supabase_client.storage.from_("image_produits").upload(
                    path=unique_filename, file=file_bytes,
                    file_options={"content-type": f"image/{file_extension.replace('.', '')}"}
                )
                final_image_url = supabase_client.storage.from_("image_produits").get_public_url(unique_filename)
            except Exception as upload_error:
                print(f"Erreur lors du téléversement de l'image : {upload_error}")
                final_image_url = DEFAULT_IMAGE

        # Détermination de la catégorie selon le switch
        category = self.p_category.value if self.p_new_category.disabled else self.p_new_category.value.strip()
        if not category:
            category = "Général"

        payload = {
            "tenant_id": self.tenant_id,
            "designation": self.p_name.value.strip(),
            "product_type": category,
            "price": float(self.p_price.value),
            "price_buy": float(self.p_price_buy.value) if self.p_price_buy.value else 0.0,
            "stock": int(self.p_stock.value or 0),
            "image": final_image_url
        }

        try:
            # 1. Premier INSERT : On laisse PostgreSQL générer l'id unique dans 'products'
            res_products = await supabase_request_async(
                access_token=self.access_token,
                tenant_id=self.tenant_id,
                table_name="products",
                method="POST",
                data=payload
            )

            # 2. Extraction sécurisée de l'ID généré pour l'alignement des tables
            generated_id = None
            if isinstance(res_products, list) and len(res_products) > 0:
                generated_id = res_products[0].get("id")
            elif isinstance(res_products, dict):
                generated_id = res_products.get("id")

            # Sécurité si l'API Supabase ne retourne pas directement la ligne créée
            if not generated_id:
                # Option de secours : aller chercher le dernier produit inséré avec cette désignation
                params_lookup = {'select': 'id', 'tenant_id': f'eq.{self.tenant_id}',
                                 'designation': f'eq.{payload["designation"]}', 'order': 'id.desc', 'limit': 1}
                lookup = await supabase_request_async(
                    access_token=self.access_token, tenant_id=self.tenant_id,
                    table_name="products", method="GET", params=params_lookup
                )
                if lookup and len(lookup) > 0:
                    generated_id = lookup[0].get("id")

            if not generated_id:
                raise Exception("Impossible de récupérer l'ID unique généré par la table products.")

            # 3. On force l'ID identique pour la table 'stock_tampons' afin d'éviter la violation de contrainte unique
            payload["id"] = int(generated_id)

            await supabase_request_async(
                access_token=self.access_token,
                tenant_id=self.tenant_id,
                table_name="stock_tampons",
                method="POST",
                data=payload
            )

            # 4. Fermeture de la modale et rechargement
            self.cp.hide_container(self.cp.st_container)

            # Pour éviter l'AssertionError Flet, on recharge les données dans le thread principal ou de façon sécurisée
            await self.load_products_data()
            self.cp.show_alert("Produit enregistré et synchronisé avec succès !", ft.Icons.CHECK_CIRCLE,
                               ft.Colors.GREEN)

        except Exception as ex:
            print(f"Erreur d'insertion du produit : {ex}")
            self.cp.show_alert(f"Erreur lors de l'enregistrement : {str(ex)}", ft.Icons.ERROR, ft.Colors.RED)

        self.loader.visible = False
        try:
            self.update()
        except Exception:
            if self.cp.page:
                self.cp.page.update()

    # --- Modification de produit ---
    def on_file_picker_result(self, e: ft.FilePickerResultEvent):
        """Déclenché lorsque l'utilisateur sélectionne une image localement"""
        if e.files and len(e.files) > 0:
            # Enregistrer le chemin local absolu de l'image
            self.selected_image_path = e.files[0].path
            # Mettre à jour l'aperçu visuel immédiat dans la modale
            self.edit_image.src = self.selected_image_path
            self.edit_image.update()
            print(f"Image locale prête pour le téléversement : {self.selected_image_path}")

    async def upload_image_to_bucket(self, file_path: str = None) -> str:
        """Téléverse l'image vers le bucket 'images' de Supabase et retourne son URL publique.
        Sécurisée avec une valeur par défaut pour éviter le bug de l'argument manquant."""
        try:
            # Si aucun chemin n'est fourni en paramètre, on récupère celui stocké dans l'instance
            path_a_charger = file_path if file_path is not None else self.selected_image_path

            if not path_a_charger or not os.path.exists(path_a_charger):
                print(f"[DEBUG] Chemin d'image invalide ou introuvable : {path_a_charger}")
                return None

            print(f"[DEBUG] Préparation de l'upload pour le fichier : {path_a_charger}")

            # Générer un nom de fichier unique pour éviter les collisions dans le bucket
            file_name = os.path.basename(path_a_charger)
            unique_name = f"{uuid.uuid4()}_{file_name}"

            # Lecture sécurisée du fichier binaire
            with open(path_a_charger, "rb") as f:
                file_data = f.read()

            # Envoi binaire vers le bucket Supabase nommé 'images'
            supabase_admin.storage.from_("images").upload(
                path=unique_name,
                file=file_data,
                file_options={"content-type": "image/jpeg"}
            )

            # Récupération de l'URL publique officielle générée par Supabase
            public_url = supabase_client.storage.from_("images").get_public_url(unique_name)
            print(f"[DEBUG] Téléversement réussi ! URL publique : {public_url}")
            return public_url

        except Exception as ex:
            print(f"[ERREUR STOCKAGE BUCKET] : {ex}")
            return None

    def open_edit_window(self, product_data: dict):
        """Ouvre la fenêtre de modification en verrouillant les prix/types"""
        # Sécurité : Intercepter 'id' ou 'produit_id' ou 'id_produit' au cas où
        self.selected_product_id = product_data.get("id") or product_data.get("produit_id") or product_data.get(
            "id_produit")

        # S'imprimer le résultat dans la console pour vérifier en temps réel
        print(f"[DEBUG] Fenêtre d'édition ouverte pour le produit : {product_data}")
        print(f"[DEBUG] ID extrait et stocké dans self.selected_product_id : {self.selected_product_id}")

        if not self.selected_product_id:
            self.cp.show_alert("Erreur interne : Impossible de récupérer l'identifiant de ce produit.", ft.Icons.ERROR,
                               ft.Colors.RED)
            return

        # 1. Remplissage des champs de texte... (le reste de ton code demeure identique)
        self.edit_designation_field.value = product_data.get("designation", "")
        self.edit_price_field.value = str(product_data.get("price", 0))
        self.edit_price_buy_field.value = str(product_data.get("price_buy", 0))
        self.edit_type_field.value = product_data.get("product_type", "")

        # 2. RESTRICTIONS STRICTES : Passage en lecture seule et désactivation des prix et types
        self.edit_price_field.read_only = True
        self.edit_price_buy_field.read_only = True
        self.edit_type_field.read_only = True

        self.edit_price_field.disabled = True
        self.edit_price_buy_field.disabled = True
        self.edit_type_field.disabled = True

        # La désignation reste totalement modifiable
        self.edit_designation_field.read_only = False
        self.edit_designation_field.disabled = False

        # 3. LOGIQUE DE L'IMAGE : Modifiable SI ET SEULEMENT SI absente ou par défaut
        current_image = product_data.get("image")
        has_prior_image = (current_image is not None and current_image != "" and current_image != DEFAULT_IMAGE)

        if has_prior_image:
            self.edit_image.src = current_image
            self.upload_btn.visible = False  # Cache le bouton d'upload si l'image existe déjà
        else:
            self.edit_image.src = DEFAULT_IMAGE
            self.upload_btn.visible = True  # Affiche le bouton d'upload si le produit n'a pas d'image

        # Réinitialisation du chemin temporaire de l'image sélectionnée
        self.selected_image_path = None

        # 4. Configuration de l'interface graphique du conteneur st_container
        self.cp.edit_ref_container.content.width = 500
        self.cp.edit_ref_container.content.height = 550
        self.cp.edit_ref_form.content.controls.clear()
        form_content = ft.Column(
            expand=True,
            spacing=15,
            controls=[
                ft.Container(
                    padding=20, bgcolor="white",
                    content=ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Text("Modifier le Produit", size=18, font_family="PEB"),
                            ft.Container(
                                content=ft.Image(resource_path("assets/icons/black/x.svg"), width=20, height=20),
                                on_click=lambda _: self.cp.hide_container(self.cp.edit_ref_container)
                            )
                        ]
                    ),
                ),
                ft.Divider(height=1, thickness=1),
                ft.Container(
                    padding=20,
                    expand=True,
                    content=ft.ListView(
                        expand=True,
                        spacing=15,
                        controls=[
                            ft.Row(
                                alignment=ft.MainAxisAlignment.CENTER,
                                controls=[
                                    ft.Stack(
                                        controls=[
                                            ft.Container(
                                                width=100, height=100, border_radius=50,
                                                border=ft.border.all(1, "grey"),
                                                content=self.edit_image
                                            ),
                                            self.upload_btn
                                        ]
                                    )
                                ]
                            ),
                            self.edit_designation_field,
                            self.edit_type_field,
                            ft.Row([self.edit_price_buy_field, self.edit_price_field], spacing=10),
                        ]
                    )
                ),
                ft.Divider(height=1, thickness=1),
                ft.Container(
                    padding=20,
                    alignment=ft.alignment.center,
                    content=MyButton(
                        "Enregistrer les modifications",
                        resource_path("assets/icons/white/check.svg"),
                        lambda e: self.run_async_in_thread(self.save_edit_product(e))
                    )
                )
            ]
        )

        self.cp.edit_ref_container.content.content = form_content
        self.cp.show_container(self.cp.edit_ref_container)
        self.cp.page.update()

    async def save_edit_product(self, e):
        """Enregistre uniquement les modifications de la désignation et de l'image
        si elle a été ajoutée avec retour visuel en cas d'erreur"""
        if not self.selected_product_id:
            self.cp.show_alert("Aucun produit sélectionné pour la modification.", ft.Icons.ERROR, ft.Colors.RED)
            return

        designation_val = self.edit_designation_field.value.strip() if self.edit_designation_field.value else ""
        if not designation_val:
            self.cp.show_alert("La désignation est obligatoire.", ft.Icons.WARNING, ft.Colors.ORANGE)
            return

        self.loader.visible = True
        self.cp.page.update()

        try:
            # Sécurité maximale : On ne construit le payload QU'AVEC la désignation au départ
            update_payload = {
                "designation": designation_val
            }

            # Si une nouvelle image a été choisie (lorsque le bouton était visible)
            if hasattr(self, 'selected_image_path') and self.selected_image_path:
                print(f"[DEBUG] Début du téléversement de l'image : {self.selected_image_path}")

                # --- CORRECTION ICI : Passage du chemin complet de l'image en paramètre ---
                url_image = await self.upload_image_to_bucket(self.selected_image_path)

                if url_image:
                    update_payload["image"] = url_image
                    print(f"[DEBUG] URL d'image récupérée avec succès : {url_image}")
                else:
                    print("[DEBUG] Échec du téléversement de l'image, l'URL est vide.")
                    self.cp.show_alert("L'image n'a pas pu être envoyée sur le serveur.", ft.Icons.WARNING,
                                       ft.Colors.ORANGE)

            print(
                f"[DEBUG] Envoi du PATCH à Supabase pour l'ID {self.selected_product_id} avec le payload : {update_payload}")

            # Mise à jour partielle (PATCH) sur Supabase
            await supabase_request_async(
                access_token=self.access_token,
                tenant_id=self.tenant_id,
                table_name="products",
                method="PATCH",
                params={"id": f"eq.{int(self.selected_product_id)}"},
                data=update_payload
            )

            print("[DEBUG] Mise à jour Supabase réussie. Rechargement des données...")

            # Fermeture du panneau et rafraîchissement complet
            self.cp.hide_container(self.cp.st_container)
            await self.load_products_data()
            self.cp.show_alert("Produit mis à jour avec succès !", ft.Icons.CHECK_CIRCLE, ft.Colors.GREEN)

        except Exception as ex:
            print(f"[ERREUR CRITIQUE] save_edit_product : {ex}")
            self.cp.show_alert(f"Erreur lors de la mise à jour : {str(ex)}", ft.Icons.ERROR, ft.Colors.RED)
        finally:
            self.loader.visible = False
            self.cp.page.update()

    # --- Import Excel ---
    def on_excel_file_picked(self, e: ft.FilePickerResultEvent):
        if not e.files:
            return
        self.run_async_in_thread(self.process_excel_import(e.files[0].path))

    async def process_excel_import(self, filepath):
        self.loader.visible = True
        self.cp.page.update()

        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active
            success_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Sécurité : Si la ligne est vide ou si la désignation (colonne 1) est absente
                if not row or row[0] is None:
                    continue

                designation = str(row[0]).strip()
                product_type = str(row[1]).strip() if row[1] else "Général"
                price_buy = float(row[2]) if row[2] is not None else 0.0
                price = float(row[3]) if row[3] is not None else 0.0
                quantite_excel = int(row[4]) if row[4] is not None else 0

                # 1. Détermination de la répartition des stocks selon le choix de l'utilisateur
                stock_principal = quantite_excel
                stock_secondaire = 0

                # 2. Préparation des deux configurations de base (sans l'id)
                payload_destination_choisie = {
                    "tenant_id": self.tenant_id,
                    "designation": designation,
                    "product_type": product_type,
                    "price_buy": price_buy,
                    "price": price,
                    "stock": stock_principal,
                    "image": DEFAULT_IMAGE
                }

                payload_destination_secondaire = {
                    "tenant_id": self.tenant_id,
                    "designation": designation,
                    "product_type": product_type,
                    "price_buy": price_buy,
                    "price": price,
                    "stock": stock_secondaire,
                    "image": DEFAULT_IMAGE
                }

                # 3. Routage dynamique des tables d'écriture
                if self.selected_stock_destination == "products":
                    table_1, payload_1 = "products", payload_destination_choisie
                    table_2, payload_2 = "stock_tampons", payload_destination_secondaire
                else:
                    table_1, payload_1 = "stock_tampons", payload_destination_choisie
                    table_2, payload_2 = "products", payload_destination_secondaire

                # 4. Premier INSERT : On laisse PostgreSQL générer l'id auto-incrémenté (serial)
                res_1 = await supabase_request_async(
                    access_token=self.access_token,
                    tenant_id=self.tenant_id,
                    table_name=table_1,
                    method="POST",
                    data=payload_1
                )

                # Extraction de l'ID généré par la première table
                generated_id = None
                if isinstance(res_1, list) and len(res_1) > 0:
                    generated_id = res_1[0].get("id")
                elif isinstance(res_1, dict):
                    generated_id = res_1.get("id")

                if not generated_id:
                    print(f"⚠️ Impossible de récupérer l'ID généré pour {designation}. Ligne ignorée.")
                    continue

                # 5. Deuxième INSERT : On force l'id pour qu'il s'aligne exactement sur le premier
                payload_2["id"] = int(generated_id)

                await supabase_request_async(
                    access_token=self.access_token,
                    tenant_id=self.tenant_id,
                    table_name=table_2,
                    method="POST",
                    data=payload_2
                )

                success_count += 1

            # Rechargement des données sur l'interface graphique
            await self.load_products_data()
            self.cp.show_alert(f"Importation réussie : {success_count} produits synchronisés !", ft.Icons.CHECK_CIRCLE,
                               ft.Colors.GREEN)

        except Exception as ex:
            print(f"Erreur d'importation Excel : {ex}")
            self.cp.show_alert("Fichier Excel invalide ou corrompu.", ft.Icons.ERROR, ft.Colors.RED)

        self.loader.visible = False
        self.cp.page.update()

    def generate_excel_template(self, e):
        """Initialise le sélecteur pour enregistrer le fichier modèle Excel structuré."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.cp.show_alert("La bibliothèque 'openpyxl' est requise pour générer le modèle.", ft.Icons.ERROR,
                               ft.Colors.RED)
            return

        def save_file_result(file_picker_event: ft.FilePickerResultEvent):
            if not file_picker_event.path:
                return

            try:
                # Création du classeur Excel
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Modèle Import Produits"

                # En-têtes attendus par ton script d'importation
                headers = ["Désignation *", "Catégorie", "Prix d'achat", "Prix de vente *", "Quantité en stock"]
                ws.append(headers)

                # Stylisation rapide des en-têtes pour faire professionnel
                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5",
                                          fill_type="solid")  # Couleur MAIN_COLOR (Indigo)
                center_alignment = Alignment(horizontal="center", vertical="center")

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                    # Ajustement de la largeur des colonnes
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 22

                # Exemple de ligne factice pour guider l'utilisateur
                ws.append(["Exemple Produit A", "Boisson", 500, 800, 15])

                # Sauvegarde à l'emplacement choisi par l'utilisateur
                wb.save(file_picker_event.path)
                self.cp.show_alert("Modèle Excel généré avec succès !", ft.Icons.CHECK_CIRCLE, ft.Colors.GREEN)

            except Exception as ex:
                print(f"Erreur lors de la création du modèle : {ex}")
                self.cp.show_alert(f"Erreur : {ex}", ft.Icons.ERROR, ft.Colors.RED)

        # Utilisation d'un FilePicker temporaire dédié à la sauvegarde
        save_picker = ft.FilePicker(on_result=save_file_result)
        self.cp.page.overlay.append(save_picker)
        self.cp.page.update()

        # Ouvre la boîte de dialogue de sauvegarde de fichier
        save_picker.save_file(
            file_name="modele_import_produits.xlsx",
            allowed_extensions=["xlsx"]
        )

    def open_import_destination_modal(self, e):
        """Ouvre l'AlertDialog de choix d'application des stocks."""
        def confirm_destination(destination_table):
            self.selected_stock_destination = destination_table
            self.import_dialog.open = False
            self.cp.page.update()
            # Déclenche l'explorateur de fichiers d'importation
            self.excel_picker.pick_files(allowed_extensions=["xlsx"])

        self.import_dialog = ft.AlertDialog(
            title=ft.Text("Destination de l'importation Excel", font_family="PEB", size=18),
            content=ft.Text(
                "Sélectionnez la table dans laquelle les quantités en stock de votre fichier Excel seront appliquées.\n\n"
                "Note : Les références seront créées simultanément dans les deux tables, mais la table secondaire démarrera avec un stock à 0.",
                size=14, font_family="PPM"
            ),
            actions=[
                ft.Row([
                    ft.ElevatedButton(
                        "Stock Boutique",
                        icon=ft.Icons.STORE_ROUNDED,
                        style=ft.ButtonStyle(color="white", bgcolor="indigo"),
                        on_click=lambda _: confirm_destination("products")
                    ),
                    ft.ElevatedButton(
                        "Stock Tampon (Dépôt)",
                        icon=ft.Icons.ALL_INBOX_ROUNDED,
                        style=ft.ButtonStyle(color="white", bgcolor="teal"),
                        on_click=lambda _: confirm_destination("stock_tampons")
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER, spacing=15)
            ],
            actions_alignment=ft.MainAxisAlignment.CENTER
        )

        self.cp.page.overlay.append(self.import_dialog)
        self.import_dialog.open = True
        self.cp.page.update()

    async def open_entree_tampon_window(self, e):
        """Ouvre la fenêtre de gestion des entrées / arrivages pour le stock tampon"""
        # Configuration des dimensions de la fenêtre de Flet (st_container ou similaire)
        self.cp.st_container.content.width = 650
        self.cp.st_container.content.height = 600

        # 1. Remplir le Dropdown avec les produits disponibles localement
        self.tampon_dropdown_produits.options.clear()
        if hasattr(self, 'liste_des_produits') and self.liste_des_produits:
            for p in self.liste_des_produits:
                p_id = str(p.get('produit_id'))
                p_nom = str(p.get('designation', 'Inconnu')).upper()
                self.tampon_dropdown_produits.options.append(
                    ft.dropdown.Option(key=p_id, text=p_nom)
                )

        # Réinitialisation du panier local et du formulaire
        self.tampon_panier_items.clear()
        self.tampon_panier_table.rows.clear()
        self.tampon_qty_field.value = ""

        # 2. Construction de la structure de l'interface
        form_content = ft.Column(
            expand=True,
            spacing=15,
            controls=[
                ft.Container(
                    padding=20, bgcolor="white",
                    content=ft.Row(
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        controls=[
                            ft.Text("Nouvel Arrivage - Stock Tampon", size=18, font_family="PEB"),
                            ft.Container(
                                content=ft.Image(resource_path("assets/icons/black/x.svg"), width=20, height=20),
                                on_click=lambda _: self.cp.hide_container(self.cp.st_container)
                            )
                        ]
                    ),
                ),
                ft.Divider(height=1, thickness=1),
                # Zone de saisie
                ft.Container(
                    padding=20,
                    content=ft.Row(
                        controls=[
                            self.tampon_dropdown_produits,
                            self.tampon_qty_field,
                            ft.IconButton(
                                icon=ft.Icons.ADD_CIRCLE_ROUNDED,
                                icon_color=MAIN_COLOR,
                                icon_size=36,
                                tooltip="Ajouter au panier",
                                on_click=self.add_item_to_tampon_panier
                            )
                        ],
                        alignment=ft.MainAxisAlignment.START,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                        spacing=10
                    )
                ),
                # Zone de la DataTable (Panier)
                ft.Container(
                    expand=True, padding=20,
                    content=ft.ListView(expand=True, controls=[self.tampon_panier_table])
                ),
                # Zone du bouton de validation au bas
                ft.Divider(height=1, thickness=1),
                ft.Container(
                    padding=20,
                    alignment=ft.alignment.center,
                    content=MyButton(
                        "Valider le panier d'entrées",
                        resource_path("assets/icons/white/check.svg"),
                        lambda e: self.run_async_in_thread(self.validate_tampon_panier(e))
                    )
                )
            ]
        )

        self.cp.st_form.content = form_content
        self.cp.show_container(self.cp.st_container)
        self.cp.page.update()

    def add_item_to_tampon_panier(self, e):
        """Ajoute temporairement un produit et sa quantité dans le tableau visuel"""
        p_id = self.tampon_dropdown_produits.value
        qty_str = self.tampon_qty_field.value

        if not p_id:
            self.cp.show_alert("Veuillez sélectionner un produit.", ft.Icons.WARNING, ft.Colors.ORANGE)
            return
        if not qty_str or not qty_str.isdigit() or int(qty_str) <= 0:
            self.cp.show_alert("Veuillez saisir une quantité positive valide.", ft.Icons.WARNING, ft.Colors.ORANGE)
            return

        qty = int(qty_str)
        # Récupérer le nom textuel du produit sélectionné dans le dropdown
        p_name = next(opt.text for opt in self.tampon_dropdown_produits.options if opt.key == p_id)

        # Vérifier si le produit est déjà présent dans le panier pour cumuler la quantité
        # Remplacer NULL par None à la fin de la fonction next()
        existing_item = next((item for item in self.tampon_panier_items if item['id_produit'] == p_id), None)
        if existing_item:
            existing_item['qte'] += qty
        else:
            self.tampon_panier_items.append({
                'id_produit': p_id,
                'designation': p_name,
                'qte': qty
            })

        # Rafraîchir la DataTable
        self.render_tampon_panier_table()

        # Réinitialiser uniquement le champ quantité pour la saisie suivante
        self.tampon_qty_field.value = ""
        self.cp.page.update()

    def render_tampon_panier_table(self):
        """Redessine les lignes de la DataTable du panier"""
        self.tampon_panier_table.rows.clear()

        for index, item in enumerate(self.tampon_panier_items):
            self.tampon_panier_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(item['designation'], font_family="PPM")),
                        ft.DataCell(ft.Text(str(item['qte']), font_family="PEB")),
                        ft.DataCell(
                            ft.Container(
                                content=ft.Image(
                                    resource_path("assets/icons/grey/trash-2.svg"), width=18, height=18,
                                ),
                                on_click=lambda e, idx=index: self.remove_item_from_tampon_panier(idx)
                            ),
                        )
                    ]
                )
            )

    def remove_item_from_tampon_panier(self, index):
        """Retire un élément du panier local"""
        if 0 <= index < len(self.tampon_panier_items):
            self.tampon_panier_items.pop(index)
            self.render_tampon_panier_table()
            self.cp.page.update()

    async def validate_tampon_panier(self, e):
        """Enregistre en bloc les entrées et met à jour la table stock_tampons"""
        if not self.tampon_panier_items:
            self.cp.show_alert("Le panier est vide.", ft.Icons.WARNING, ft.Colors.ORANGE)
            return

        self.loader.visible = True
        self.cp.page.update()

        try:
            for item in self.tampon_panier_items:
                prod_id = int(item['id_produit'])
                quantite_entree = item['qte']

                # 1. Récupérer le stock tampon actuel du produit pour faire le cumul exact
                params = {'select': 'stock', 'id': f'eq.{prod_id}'}
                res_tampon = await supabase_request_async(
                    access_token=self.access_token, tenant_id=self.tenant_id,
                    table_name="stock_tampons", method="GET", params=params
                )

                current_stock_tampon = 0
                if res_tampon and len(res_tampon) > 0:
                    current_stock_tampon = int(res_tampon[0].get('stock') or 0)

                # 2. Calculer et appliquer le nouveau stock tampon
                new_stock_tampon = current_stock_tampon + quantite_entree

                await supabase_request_async(
                    access_token=self.access_token, tenant_id=self.tenant_id,
                    table_name="stock_tampons", method="PATCH",
                    params={"id": f"eq.{prod_id}"},
                    data={"stock": new_stock_tampon}
                )

                # 3. Insérer la ligne d'historique correspondante dans la nouvelle table entrees_tampons
                payload_entree_tampon = {
                    "tenant_id": self.tenant_id,
                    "id_produit": prod_id,
                    "qte": quantite_entree,
                    "cree_par": self.user_name.upper() if hasattr(self, 'user_name') else "ADMIN"
                }

                await supabase_request_async(
                    access_token=self.access_token, tenant_id=self.tenant_id,
                    table_name="entrees_tampons", method="POST",
                    data=payload_entree_tampon
                )

            # 4. Succès total et rechargement de la vue générale
            self.cp.hide_container(self.cp.st_container)
            await self.load_products_data()  # Recharge tes tableaux de bord principaux
            self.cp.show_alert("Arrivage en stock tampon enregistré avec succès !", ft.Icons.CHECK_CIRCLE,
                               ft.Colors.GREEN)

        except Exception as ex:
            print(f"Erreur validation panier tampon : {ex}")
            self.cp.show_alert("Une erreur est survenue durant l'enregistrement.", ft.Icons.ERROR, ft.Colors.RED)
        finally:
            self.loader.visible = False
            self.cp.page.update()


